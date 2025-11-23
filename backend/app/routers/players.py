from typing import List, Optional, Sequence, Tuple
import csv
import io
import math
from pathlib import Path

from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, Path
from sqlalchemy.orm import Session

from ..db import get_db
from .. import models, schemas
from ..security import get_current_agent

try:  # pragma: no cover - optional dependency handled at runtime
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - handled in endpoint
    load_workbook = None


router = APIRouter(prefix="/players", tags=["players"]) #endpoint /players
MAX_BULK_IMPORT_ROWS = 200 # to prevent abuse


SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SUPPORTED_CSV_SUFFIXES = {".csv", ".txt"}
FULL_NAME_HEADERS = {"full_name", "fullname", "name", "player", "player_name"}
PHONE_HEADERS = {"phone_number", "phonenumber", "phone", "mobile", "phone_no", "contact"}


def _detect_file_format(file: UploadFile) -> str:
    suffix = Path(file.filename or "").suffix.lower()
    content_type = (file.content_type or "").lower()

    if suffix in SUPPORTED_CSV_SUFFIXES or "csv" in content_type or content_type.startswith("text/"):
        return "csv"

    if suffix in SUPPORTED_EXCEL_SUFFIXES or content_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    }:
        return "excel"

    raise HTTPException(
        status_code=400,
        detail="Unsupported file type. Please upload a CSV or Excel (.xlsx) file.",
    )


def _read_csv_rows(data: bytes) -> List[Tuple[int, List[str]]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise HTTPException(
            status_code=400,
            detail="Could not decode the uploaded file. Please make sure it is a valid UTF-8 encoded CSV file.",
        )

    reader = csv.reader(io.StringIO(text))
    rows: List[Tuple[int, List[str]]] = []
    for row_number, row in enumerate(reader, start=1):
        if not any(cell.strip() for cell in row):
            continue
        rows.append((row_number, row))
    return rows


def _read_excel_rows(data: bytes) -> List[Tuple[int, List[Optional[object]]]]:
    if load_workbook is None:
        raise HTTPException(
            status_code=400,
            detail="Excel support is unavailable on the server.",
        )

    stream = io.BytesIO(data)
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Unable to read the Excel file. Please upload a valid .xlsx workbook.",
        )

    worksheet = workbook.active
    rows: List[Tuple[int, List[Optional[object]]]] = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        rows.append((row_number, list(row)))
    return rows


def _normalise_full_name(value: Optional[object]) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalise_phone_number(value: Optional[object]) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return None
        if value.is_integer():
            return str(int(value))
        return str(value).strip()

    phone = str(value).strip()
    if not phone:
        return None
    if phone.lower() in {"nan", "none", "null"}:
        return None
    return phone


def _looks_like_header_row(full_name: str, phone_number: Optional[str]) -> bool:
    name_key = full_name.strip().lower().replace(" ", "_")
    if name_key not in FULL_NAME_HEADERS:
        return False

    if phone_number is None:
        return True

    phone_key = phone_number.strip().lower().replace(" ", "_")
    return phone_key in PHONE_HEADERS or phone_key == ""


def _extract_player_rows(rows: Sequence[Tuple[int, Sequence[Optional[object]]]]) -> List[Tuple[int, str, Optional[str]]]:
    extracted: List[Tuple[int, str, Optional[str]]] = []
    full_idx = 0
    phone_idx = 1
    header_detected = False

    iterator = iter(rows)
    try:
        first_row_number, first_values = next(iterator)
    except StopIteration:
        return extracted

    normalised_header = [
        str(value).strip().lower().replace(" ", "_") if value is not None else ""
        for value in first_values
    ]

    detected_full_idx: Optional[int] = None
    detected_phone_idx: Optional[int] = None
    for idx, value in enumerate(normalised_header):
        if detected_full_idx is None and value in FULL_NAME_HEADERS:
            detected_full_idx = idx
        if detected_phone_idx is None and value in PHONE_HEADERS:
            detected_phone_idx = idx

    if detected_full_idx is not None and detected_phone_idx is not None:
        full_idx = detected_full_idx
        phone_idx = detected_phone_idx
        header_detected = True
    else:
        # Treat the first row as data
        iterator = iter(rows)

    if not header_detected:
        first_row_number, first_values = next(iterator)
        extracted.append(
            (
                first_row_number,
                _normalise_full_name(first_values[full_idx] if len(first_values) > full_idx else None),
                _normalise_phone_number(first_values[phone_idx] if len(first_values) > phone_idx else None),
            )
        )

    for row_number, values in iterator:
        full_value = values[full_idx] if len(values) > full_idx else None
        phone_value = values[phone_idx] if len(values) > phone_idx else None
        extracted.append(
            (
                row_number,
                _normalise_full_name(full_value),
                _normalise_phone_number(phone_value),
            )
        )

    return extracted


def _parse_player_rows(file: UploadFile, raw: bytes) -> List[Tuple[int, str, Optional[str]]]:
    file_format = _detect_file_format(file)
    if file_format == "csv":
        base_rows = _read_csv_rows(raw)
    else:
        base_rows = _read_excel_rows(raw)

    if not base_rows:
        return []

    try:
        players = _extract_player_rows(base_rows)
    except StopIteration:
        return []

    return players


@router.get("", response_model=List[schemas.PlayerOut])
def list_players(
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    return (
        db.query(models.Player)
        .filter(models.Player.agent_id == current_agent.id)
        .order_by(models.Player.created_at.desc())
        .all()
    ) #list all players ordered by created_at desc

@router.get("/{phone_number}", response_model=schemas.PlayerOut) #get player by phone number
def get_player(
    phone_number: str,
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    player = (
        db.query(models.Player)
        .filter(
            models.Player.phone_number == phone_number,
            models.Player.agent_id == current_agent.id,
        )
        .first()
    )
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    return player


@router.post("", response_model=schemas.PlayerOut, status_code=201) #create a new player
def create_player(
    payload: schemas.PlayerCreate,
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    existsing_player = (
        db.query(models.Player)
        .filter(
            models.Player.phone_number == payload.phone_number,
            models.Player.agent_id == current_agent.id,
        )
        .first()
    )
    if existsing_player:
        raise HTTPException(status_code=400 , detail="This phone number already exists in the Player database")
    player=models.Player(
        agent_id=current_agent.id,
        full_name=payload.full_name,
        phone_number=payload.phone_number
    )
    db.add(player)
    db.commit()
    db.refresh(player)
    return player

async def _bulk_import_players(
    file: UploadFile,
    db: Session,
    current_agent: models.Agent,
) -> schemas.BulkImportResult:
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    rows = _parse_player_rows(file, raw)
    if not rows:
        raise HTTPException(status_code=400, detail="No player rows found in the uploaded file.")

    existing_phone_numbers = {
        str(pn).strip()
        for (pn,) in db.query(models.Player.phone_number)
        .filter(models.Player.agent_id == current_agent.id)
        .all()
        if pn
    }

    created_count = 0
    processed = 0
    errors: List[str] = []

    for idx, (row_number, full_name, phone_number) in enumerate(rows, start=1):
        if processed >= MAX_BULK_IMPORT_ROWS:
            errors.append(
                f"Maximum limit of {MAX_BULK_IMPORT_ROWS} rows reached. Some rows were not processed."
            )
            break

        if idx == 1 and _looks_like_header_row(full_name, phone_number if isinstance(phone_number, str) else None):
            continue

        processed += 1

        if not full_name:
            errors.append(f"Row {row_number}: full_name is required.")
            continue

        phone_value = phone_number.strip() if isinstance(phone_number, str) else phone_number
        if phone_value and phone_value in existing_phone_numbers:
            errors.append(f"Row {row_number}: phone_number '{phone_value}' already exists.")
            continue

        player = models.Player(
            agent_id=current_agent.id,
            full_name=full_name,
            phone_number=phone_value,
        )
        db.add(player)
        try:
            db.commit()
            db.refresh(player)
            created_count += 1
            if phone_value:
                existing_phone_numbers.add(phone_value)
        except Exception as e:
            db.rollback()
            errors.append(f"Row {row_number}: Database error: {str(e)}")

    return schemas.BulkImportResult(
        total_rows=processed,
        created=created_count,
        skipped=processed - created_count,
        errors=errors,
    )


@router.post("/import", response_model=schemas.BulkImportResult)
async def import_players(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    return await _bulk_import_players(file, db, current_agent)


@router.post("/import-csv", response_model=schemas.BulkImportResult)
async def import_players_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    """Backward compatible endpoint that shares the same importer."""
    return await _bulk_import_players(file, db, current_agent)

#------------Alter Player----------------
#alter phone number
@router.put("/{phone_number}", response_model=schemas.PlayerOut)
def update_player(
    phone_number: str,
    payload: schemas.PlayerCreate,
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    player = (
        db.query(models.Player)
        .filter(
            models.Player.phone_number == phone_number,
            models.Player.agent_id == current_agent.id,
        )
        .first()
    )
    if not player:
        raise HTTPException(status_code=404, detail="Player not found with that number")
    
    if payload.phone_number != phone_number:
        # Check for uniqueness of new phone number
        existing_player = (
            db.query(models.Player)
            .filter(
                models.Player.phone_number == payload.phone_number,
                models.Player.agent_id == current_agent.id,
            )
            .first()
        )
        if existing_player:
            raise HTTPException(status_code=400, detail="This new phone number already exists in the Player database")
    
    player.phone_number = payload.phone_number
    db.commit()
    db.refresh(player)
    return player

#------------DELETE PLAYER/S----------------
@router.delete("/id/{player_id}", status_code=204)
def delete_player_by_id(
    player_id: int,
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    player = (
        db.query(models.Player)
        .filter(models.Player.id == player_id, models.Player.agent_id == current_agent.id)
        .first()
    )
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    db.delete(player)
    db.commit()
    return None


@router.delete("/{phone_number}", status_code=204)
def delete_player(
    phone_number: str,
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    player = (
        db.query(models.Player)
        .filter(
            models.Player.phone_number == phone_number,
            models.Player.agent_id == current_agent.id,
        )
        .first()
    )
    if not player:
        raise HTTPException(status_code=404, detail="Player not found with that number")
    db.delete(player)
    db.commit()
    return None

@router.delete("", status_code=204)
def delete_all_players(
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    (
        db.query(models.Player)
        .filter(models.Player.agent_id == current_agent.id)
        .delete(synchronize_session=False)
    )
    db.commit()
    return


#------------Get Player State----------------  
@router.get("/state/{event_id}/{player_id}", response_model=schemas.PlayerStateOut)
def player_state_by_id(
    event_id: int = Path(...),
    player_id: int = Path(...),
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    _get_event_or_404(db, event_id, current_agent.id)
    p = (
        db.query(models.Player)
        .filter(models.Player.id == player_id, models.Player.agent_id == current_agent.id)
        .first()
    )
    if not p:
        raise HTTPException(status_code=404, detail="Player not found")
    return _player_state_payload(db, event_id, p)

@router.get("/state/by-phone/{event_id}/{phone_number}", response_model=schemas.PlayerStateOut)
def player_state_by_phone(
    event_id: int = Path(...),
    phone_number: str = Path(...),
    db: Session = Depends(get_db),
    current_agent: models.Agent = Depends(get_current_agent),
):
    _get_event_or_404(db, event_id, current_agent.id)
    p = (
        db.query(models.Player)
        .filter(
            models.Player.phone_number == phone_number,
            models.Player.agent_id == current_agent.id,
        )
        .first()
    )
    if not p:
        raise HTTPException(status_code=404, detail="Player not found")
    return _player_state_payload(db, event_id, p)





#------------Helper Functions----------------
def _get_event_or_404(db: Session, event_id: int, agent_id: int) -> models.Event:
    ev = (
        db.query(models.Event)
        .filter(models.Event.id == event_id, models.Event.agent_id == agent_id)
        .first()
    )
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    return ev

def _player_state_payload(
    db: Session, event_id: int, player: models.Player
) -> schemas.PlayerStateOut:
    # Find an ACTIVE assignment involving this player in this event
    a = (
        db.query(models.Assignment)
        .filter(
            models.Assignment.event_id == event_id,
            models.Assignment.status == "active",
            or_(
                models.Assignment.player1_id == player.id,
                models.Assignment.player2_id == player.id,
            ),
        )
        .first()
    )
    if not a:
        return schemas.PlayerStateOut(
            player=player,
            state="free",
            assignment_id=None,
            table_id=None,
            table_position=None,
            opponent=None,
        )

    # Get table label if available
    table = None
    if a.table_id:
        table = db.query(models.Table).filter(models.Table.id == a.table_id).first()

    # Determine opponent
    opponent_id = a.player2_id if a.player1_id == player.id else a.player1_id
    opponent = db.query(models.Player).filter(models.Player.id == opponent_id).first()

    return schemas.PlayerStateOut(
        player=player,
        state="playing",
        assignment_id=a.id,
        table_id=a.table_id,
        table_position=table.position if table else None,
        opponent=opponent,
    )